"""The Croatian licensing register, as a corroboration source.

The brief names Destination Management Organisations explicitly, so a second,
non-Google source was always in scope. What it turned into is more useful than
a second discovery stream.

WHAT IT IS
    The Ministry of Tourism publishes *Popis turističkih agencija*, every travel
    agency licensed to trade in Croatia. One XLSX, ~1,750 rows, refreshed
    weekly, served without a key or a rate limit under an explicit open licence
    ("Otvorena dozvola", commercial reuse permitted). Neither host publishes a
    robots.txt. This is a published dataset being downloaded, not a site being
    scraped, which matters given the no-scraping position taken everywhere else.

WHY IT IS NOT A DISCOVERY SOURCE
    It looked like one. 210 of the 215 agencies with Split premises have no
    name counterpart among the 301 places Google returned, which reads as a
    coverage gap worth 210 leads. It is not. The register is a LEGAL CATEGORY,
    not a trade: spot-checking the unmatched entries found a villa-rental
    business, a financial services company, an import/export trader, a property
    consultancy, and several with no findable trading presence at all. Claiming
    those as net-new operators would be the same error as claiming a precision
    figure from synthetic data.

    So the yield gets measured before the source gets wired. Until then this
    module deliberately does not implement the DiscoverySource protocol, and
    does not emit DiscoveredPlace, because a register row is not a place: it
    carries no coordinates, no phone, no website and no rating.

WHAT IT IS GOOD FOR, AND THE ANSWER IS: LESS THAN IT LOOKS
    The plan was to use it for licensing facts about operators already
    discovered. Whether an operator is licensed to sell travel, carries
    liability insurance and holds insolvency protection are hard regulatory
    signals about readiness to transact on a marketplace, and everything else
    on that axis is inferred from a website.

    Measured, that plan fails, and it fails in an informative way:

        corroborated licence matches   5 of 167 operators   (3.0%)
        ... of the 40 published leads  0                    (0.0%)

    The floor is not what is doing this. The median best-match score across all
    167 operators is 0.18, and 70 of them share no distinctive word with ANY
    licensed agency in Split. Loosening the corroboration floor from 0.70 to
    0.50 moves the join from 6 to 15, which is noise, not coverage.

    The five it does match are Gray Line Croatia, Croatia Private Tours,
    Adriatic Vision and two others of the same kind: established agencies. Those
    are precisely the operators most likely to be on the marketplace already,
    which is why none of them is a lead. The register vouches for the supply
    Viator already has and is blind to the supply it wants.

    The reason is that "turistička agencija" is a legal category, not a trade. A
    skipper running boat excursions is regulated as maritime transport, a guide
    as a guide, and many operators sell through an agency rather than being one.

    So the register is NOT wired into scoring and NOT attached to the lead
    record. A signal observable for 3% of the population cannot carry a scoring
    axis without introducing exactly the bias the review-volume cap and the
    rating prior exist to counteract, and attaching a field that is empty on all
    40 leads would be decoration.

    It stays in the repo because the measurement is the point, and because the
    multi-source architecture is now demonstrated by a real second source rather
    than asserted by a Protocol with one implementation. Run
    scripts/registry_check.py to reproduce every number above.

HOW IT MUST BE JOINED
    Not on address alone. Measured: a street-address join links 49 of 248
    keyed places, and most of those links are wrong. Obala Lazareta 3 returns
    five different Google businesses and one register entry; Trumbićeva obala 2
    and 13 do the same. That stretch of waterfront is a row of booking kiosks,
    so the address identifies a berth, not a business.

    This is the same finding as the shared phone (13.2% of operators) and the
    shared host (eight operators on one wixsite subdomain), now for the third
    time in a third signal, in a source added after the pattern was written.
    The established rule applies unchanged: corroborate with the name, or do
    not decide.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import httpx
import openpyxl

from supply_radar.locales import LocalePack
from supply_radar.matching import name_similarity
from supply_radar.normalise import fold_diacritics, normalise_name

INDEX_URL = (
    "https://appl.mint.hr/upisnik-turistickih-agencija/"
    "popis-turistickih-agencija-od-1-1-2018/8"
)
USER_AGENT = "SupplyRadar/0.1 (case study prototype; contact via repo)"

# Filenames on the index are inconsistent enough that constructing one from
# today's date does not work: 250307_Popis_TA..xlsx carries a double dot. The
# latest file is found by parsing the index, never by building a URL.
_HREF = re.compile(r'href="([^"]*?(\d{6})_Popis_TA\.*\.xlsx[^"]*)"', re.IGNORECASE)

# The premises column is "street and number, postcode town | County", and a
# single agency may list several premises on separate lines.
_PREMISES_COUNTY = re.compile(r"\|")
_POSTCODE_TOWN = re.compile(r"\b(\d{2}\s?\d{3})\s+(.+?)\s*$")

# County spellings vary in the published file. Folding diacritics is not enough
# because two of them are genuinely misspelt at source.
_COUNTY_ALIASES = {
    "osiecko-baranjska": "osjecko-baranjska",
    "sisacko-moslovacka": "sisacko-moslavacka",
    "dubrovacko - neretvanska": "dubrovacko-neretvanska",
    "primorsko -goranska": "primorsko-goranska",
    "istra": "istarska",
    "zadar": "zadarska",
    "pozega": "pozesko-slavonska",
}

# The register answers DA (yes) or NE (no), in mixed case, and leaves the cell
# blank or "-" when the question was never answered. Blank is not "no": it means
# the ministry holds no declaration, which is a different fact and is kept
# distinct rather than flattened into False.
_YES = {"da"}
_NO = {"ne"}


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _tristate(value) -> bool | None:
    text = (_clean(value) or "").lower()
    if text in _YES:
        return True
    if text in _NO:
        return False
    return None


# Register names put the legal form in the MIDDLE, not at the end:
# "WATERWORLD j.d.o.o. za turizam i usluge, turistička agencija". normalise_name
# strips a trailing legal suffix, which is right for every other source here
# (Google returns trading names, and the synthetic CRM appends its legal forms),
# but leaves "jdoo" sitting in the middle of a register name looking like a
# distinctive word. Measured, that alone dragged Waterworld Holidays against
# WATERWORLD j.d.o.o. down to 0.555, under the corroboration floor.
#
# Scoped to this module rather than the locale pack because it is a property of
# how this register writes names, not of Croatian, and the core normaliser is
# load-bearing for the published matching numbers.
_LEGAL_FORM_ANYWHERE = re.compile(
    r"(?<!\w)(j\.?\s?d\.?\s?o\.?\s?o\.?|d\.?\s?o\.?\s?o\.?|d\.?\s?d\.?"
    r"|j\.?\s?t\.?\s?d\.?|k\.?\s?d\.?)(?!\w)",
    re.IGNORECASE,
)


def register_name(legal_name: str, locale: LocalePack) -> str:
    """The comparable identity key for a register entry."""
    return normalise_name(_LEGAL_FORM_ANYWHERE.sub(" ", legal_name), locale)


def _slug(text: str) -> str:
    folded = unicodedata.normalize("NFKD", text.lower())
    folded = "".join(c for c in folded if not unicodedata.combining(c))
    folded = folded.replace("đ", "d")
    return re.sub(r"\s+", " ", folded).strip()


@dataclass(frozen=True)
class RegisterEntry:
    """One licensed agency at one set of premises.

    An agency operating from three addresses produces three entries, because
    the premises is what a destination-scoped search needs to match on.
    """

    legal_name: str
    registered_office: str | None
    premises: str
    town: str | None
    county: str | None
    liability_insurance: bool | None
    insolvency_protection: bool | None
    insurer: str | None
    policy_expiry: str | None
    first_registered: str | None

    @property
    def is_fully_protected(self) -> bool:
        """Both protections declared. The state a marketplace wants to see."""
        return bool(self.liability_insurance and self.insolvency_protection)


# ------------------------------------------------------------------ fetching


def latest_file_url(client: httpx.Client | None = None) -> tuple[str, str]:
    """Return (url, yymmdd) for the most recent published register file."""
    owns = client is None
    client = client or httpx.Client(
        timeout=60, headers={"User-Agent": USER_AGENT}, follow_redirects=True
    )
    try:
        response = client.get(INDEX_URL)
        response.raise_for_status()
        found = _HREF.findall(response.text)
    finally:
        if owns:
            client.close()

    if not found:
        raise RuntimeError(
            f"No Popis_TA .xlsx link found on {INDEX_URL}. The ministry has "
            "changed the page layout; the parser needs updating."
        )
    url, stamp = max(found, key=lambda pair: pair[1])
    return url, stamp


def fetch_register(cache_dir: Path, refresh: bool = False) -> Path:
    """Download the latest register, or reuse the cached copy.

    Cached by the file's own published date rather than by fetch time, so a
    re-run on the same day is free and a re-run after the weekly refresh picks
    the new file up on its own.
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    with httpx.Client(
        timeout=60, headers={"User-Agent": USER_AGENT}, follow_redirects=True
    ) as client:
        url, stamp = latest_file_url(client)
        target = cache_dir / f"{stamp}_popis_ta.xlsx"
        if target.exists() and not refresh:
            return target
        response = client.get(url)
        response.raise_for_status()
        target.write_bytes(response.content)
    return target


# ------------------------------------------------------------------- parsing


def parse_register(path: Path) -> list[RegisterEntry]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    entries: list[RegisterEntry] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue
        legal_name = _clean(row[2])
        if not legal_name:
            continue

        premises_cell = _clean(row[4]) or _clean(row[3]) or ""
        for line in premises_cell.split("\n"):
            premises = line.strip()
            if not premises:
                continue
            address, county = _split_county(premises)
            entries.append(
                RegisterEntry(
                    legal_name=legal_name,
                    registered_office=_clean(row[3]),
                    premises=address,
                    town=_town(address),
                    county=county,
                    liability_insurance=_tristate(row[5]),
                    insurer=_clean(row[6]),
                    policy_expiry=_clean(row[8]),
                    insolvency_protection=_tristate(row[9]),
                    first_registered=_clean(row[1]),
                )
            )
    return entries


def _split_county(premises: str) -> tuple[str, str | None]:
    if "|" not in premises:
        return premises.strip(), None
    address, _, county = premises.rpartition("|")
    slug = _slug(county)
    slug = slug.rstrip(",").strip()
    return address.strip(), _COUNTY_ALIASES.get(slug, slug) or None


def _town(address: str) -> str | None:
    match = _POSTCODE_TOWN.search(address.strip())
    if not match:
        return None
    return _slug(match.group(2))


def entries_for_town(entries: list[RegisterEntry], town: str) -> list[RegisterEntry]:
    """Entries whose PREMISES are in this town.

    Matching on the town parsed out of the address, never on a substring of the
    whole line: "Splitsko-dalmatinska" is a county containing the string
    "Split", and every agency in the county would otherwise be counted as
    trading in the city.
    """
    target = _slug(town)
    return [e for e in entries if e.town == target]


# --------------------------------------------------------------- the join


@dataclass(frozen=True)
class LicenceMatch:
    entry: RegisterEntry
    name_agreement: float
    address_agreement: bool


# Corroboration floor. Deliberately the same 0.70 the phone and domain hard keys
# use: this is the same rule about the same class of signal, and giving it a
# different number would imply a distinction that does not exist.
NAME_CORROBORATION_FLOOR = 0.70

_STREET_TYPES = {"ulica", "obala", "put", "trg", "cesta", "setaliste", "kneza", "ul"}
_HOUSE_NUMBER = re.compile(r"\b(\d+)\s*[a-z]?\b")


def street_key(address: str | None, locale: LocalePack) -> str | None:
    """A comparable "this street, this number" key.

    Croatian addresses carry suffixes the two sources disagree about ("12a",
    "10 / V kat"), so the number is reduced to its leading digits and the
    street to its most distinctive word.
    """
    if not address:
        return None
    folded = fold_diacritics(address.lower(), locale).split("|")[0]
    folded = re.sub(r"\b\d{2}\s?\d{3}\b", " ", folded)
    match = _HOUSE_NUMBER.search(folded)
    if not match:
        return None
    words = [w for w in re.split(r"[^a-z]+", folded[: match.start()]) if len(w) > 2]
    words = [w for w in words if w not in _STREET_TYPES]
    if not words:
        return None
    return f"{words[-1]}|{match.group(1)}"


def find_licence(
    name: str,
    address: str | None,
    entries: list[RegisterEntry],
    locale: LocalePack,
    idf=None,
) -> LicenceMatch | None:
    """The register entry for this operator, or None.

    Address narrows the field and the name decides. Address alone is not
    allowed to decide, for the reason recorded at the top of this module: on
    the Split waterfront one address routinely covers five businesses.
    """
    if not entries:
        return None

    key = street_key(address, locale)
    candidates = (
        [e for e in entries if street_key(e.premises, locale) == key] if key else []
    )
    # Falling back to the whole set when the address does not parse is
    # deliberate. The name still has to clear the floor on its own, so this
    # widens the search without weakening the decision.
    pool = candidates or entries

    place_norm = normalise_name(name, locale)
    if not place_norm:
        return None

    best: LicenceMatch | None = None
    for entry in pool:
        entry_norm = register_name(entry.legal_name, locale)
        if not entry_norm:
            continue
        score, _ = name_similarity(place_norm, entry_norm, idf)
        if score >= NAME_CORROBORATION_FLOOR and (
            best is None or score > best.name_agreement
        ):
            best = LicenceMatch(
                entry=entry,
                name_agreement=score,
                address_agreement=entry in candidates,
            )
    return best
